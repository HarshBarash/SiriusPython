import warnings

from openpyxl import load_workbook
wb_males = load_workbook("/home/barash/PycharmProjects/Sirius/data-6271-2020-10-16.xlsx")
wb_females = load_workbook("/home/barash/PycharmProjects/Sirius/data-6271-2020-10-16.xlsx")
sheet_male = wb_males['Sheet0']
sheet_female = wb_females['Sheet0']


name = input()
while name != "end":
    for i in range(2, 6801):
        d = sheet_male.cell(column=2, row=i).value
        f = sheet_female.cell(column=2, row=i).value
        if d == name:
            print("Уважаемый", name)
            break
        if f == name:
            print("Уважаемая", name)
            break






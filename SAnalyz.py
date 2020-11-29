from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.ticker as tick
import tkinter
import matplotlib
matplotlib.use('TkAgg')

alphabet = ['а', 'б', 'в', 'г', 'д', 'е', 'и',
            'й', 'к', 'л', 'м', 'н', 'п', 'р',
            'с', 'т', 'у', 'ф', 'х', 'ь', 'я']
wb_males = load_workbook("./data-6271-2020-10-16.xlsx")
wb_females = load_workbook("./data-6269-2020-10-16.xlsx")
counters_males = [0]*21
counters_females = [0]*21
sheet_male = wb_males['Sheet0']
sheet_female = wb_females['Sheet0']
count_male = 0
count_female = 0
fig, ax = plt.subplots(figsize=(6, 11))



for i in alphabet:
    for j in range(2, 6801):
        d = sheet_male.cell(column=2, row=j).value
        f = sheet_female.cell(column=2, row=j).value
        if d[-1] == i:
            counters_males[count_male] += \
                int(sheet_male.cell(column=3, row=j).value)
        if f[-1] == i:
            counters_females[count_female] += \
                int(sheet_female.cell(column=3, row=j).value)
    count_male += 1
    count_female += 1


plt.title('Dependence of the last letter on the name')
plt.xlabel('Last letter of the name')
plt.ylabel('Number of births')
plt.xticks(range(len(alphabet)), alphabet)
ax.yaxis.grid(True)
ax.yaxis.set_major_locator(tick.MultipleLocator(8000))
plt.plot(alphabet, counters_males, color='grey',
         linestyle='solid', label='Males')
plt.plot(alphabet, counters_females, color='purple',
         linestyle='solid', label='Females')
plt.legend(loc='upper center')
plt.show()

